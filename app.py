"""
Trendomat by Odyseo — narzędzie do analizy trendów wyszukiwań dla dowolnego klienta/marki.

Sprawdza dla wklejonej listy fraz: wolumen wyszukiwań (aktualny i sprzed roku) + Google Trends,
a opcjonalnie także wolumen wyszukiwań fraz brandowych konkurencji (bez Google Trends).

Źródła danych:
- DataForSEO: keywords_data/google_ads/search_volume/live -> wolumen aktualny + historia miesięczna
              (z historii wyciągamy wartość sprzed ~12 miesięcy jako "rok temu")
- DataForSEO: keywords_data/google_trends/explore/live -> szereg czasowy trendu (0-100) dla frazy
- Senuto: integracja "best effort" — Senuto nie ma w pełni publicznej, ujednoliconej
          dokumentacji parametrów dla konta czytelnika. Uzupełnij funkcję `senuto_get_volumes`
          zgodnie z endpointem dostępnym na Twoim planie (docs-api.senuto.com), jeśli chcesz
          traktować Senuto jako źródło danych PL zamiast/obok DataForSEO.

Uruchomienie lokalnie:
    pip install -r requirements.txt
    streamlit run app.py
"""

import base64
import datetime as dt
import io
import time

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import requests
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (Image, KeepTogether, PageBreak, Paragraph,
                                 SimpleDocTemplate, Spacer, Table, TableStyle)

st.set_page_config(page_title="Trendomat by Odyseo", layout="wide")

DFS_BASE = "https://api.dataforseo.com/v3"
SENUTO_BASE = "https://api.senuto.com/api"

# Kolorystyka Odyseo: czarne tło, biel i fiolet marki jako akcent.
ODYSEO_PURPLE = "#422AAB"
ODYSEO_PURPLE_LIGHT = "#8B6FE8"  # jaśniejszy fiolet do linii/tekstu na czarnym tle
PDF_BLACK = colors.HexColor("#000000")
PDF_WHITE = colors.white
PDF_PURPLE = colors.HexColor(ODYSEO_PURPLE)
PDF_PURPLE_LIGHT = colors.HexColor(ODYSEO_PURPLE_LIGHT)
PDF_GRID = colors.HexColor("#3A3A3A")
PDF_GREEN = colors.HexColor("#22C55E")
PDF_RED = colors.HexColor("#EF4444")

# Font z polskimi znakami — DejaVu Sans jest dołączony do matplotlib, więc nie trzeba
# dorzucać osobnego pliku .ttf do repozytorium.
_MPL_FONT_DIR = matplotlib.get_data_path() + "/fonts/ttf"
pdfmetrics.registerFont(TTFont("DejaVuSans", _MPL_FONT_DIR + "/DejaVuSans.ttf"))
pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", _MPL_FONT_DIR + "/DejaVuSans-Bold.ttf"))

# ---------------------------------------------------------------------------
# Pomocnicze
# ---------------------------------------------------------------------------

def parse_phrases(raw_text: str) -> list[str]:
    phrases = [p.strip() for p in raw_text.splitlines() if p.strip()]
    # usuń duplikaty, zachowaj kolejność
    seen = set()
    result = []
    for p in phrases:
        key = p.lower()
        if key not in seen:
            seen.add(key)
            result.append(p)
    return result


def chunked(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = (text.replace("ą", "a").replace("ć", "c").replace("ę", "e").replace("ł", "l")
            .replace("ń", "n").replace("ó", "o").replace("ś", "s").replace("ź", "z").replace("ż", "z"))
    out = "".join(c if c.isalnum() else "_" for c in text)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_")


# ---------------------------------------------------------------------------
# DataForSEO
# ---------------------------------------------------------------------------

def dfs_auth_header(login: str, password: str) -> dict:
    token = base64.b64encode(f"{login}:{password}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}


def dfs_search_volume(login: str, password: str, keywords: list[str], location_code: int, language_code: str) -> pd.DataFrame:
    """Aktualny wolumen + odczyt sprzed 12 miesięcy z monthly_searches."""
    headers = dfs_auth_header(login, password)
    rows = []
    for batch in chunked(keywords, 700):  # limit 1000/req, zapas na bezpieczeństwo
        payload = [{
            "keywords": batch,
            "location_code": location_code,
            "language_code": language_code,
            "search_partners": False,
        }]
        resp = requests.post(f"{DFS_BASE}/keywords_data/google_ads/search_volume/live",
                              headers=headers, json=payload, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        tasks = data.get("tasks", [])
        for task in tasks:
            for item in (task.get("result") or []):
                if item is None:
                    continue
                kw = item.get("keyword")
                current_vol = item.get("search_volume")
                monthly = item.get("monthly_searches") or []
                # monthly_searches: lista {"year":..,"month":..,"search_volume":..}, ostatnie 12 mies.
                vol_year_ago = None
                if monthly:
                    monthly_sorted = sorted(monthly, key=lambda m: (m["year"], m["month"]))
                    if len(monthly_sorted) >= 12:
                        vol_year_ago = monthly_sorted[-12]["search_volume"]
                    else:
                        vol_year_ago = monthly_sorted[0]["search_volume"]
                rows.append({
                    "fraza": kw,
                    "wolumen_aktualny": current_vol,
                    "wolumen_rok_temu": vol_year_ago,
                    "cpc": item.get("cpc"),
                    "konkurencja": item.get("competition"),
                })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["zmiana_%"] = df.apply(
            lambda r: round(((r["wolumen_aktualny"] - r["wolumen_rok_temu"]) / r["wolumen_rok_temu"] * 100), 1)
            if r["wolumen_rok_temu"] not in (None, 0) and r["wolumen_aktualny"] is not None else None,
            axis=1,
        )
    return df


def dfs_trends_explore(login: str, password: str, keywords: list[str], location_code: int,
                        time_range: str = "past_5_years", max_retries: int = 2,
                        progress_callback=None) -> tuple[dict[str, pd.DataFrame], list[str], dict[str, str]]:
    """Zwraca (słownik fraza -> DataFrame(date, interest), lista fraz które się nie udały, powody).

    WAŻNE: endpoint `google_trends/explore/live` w DataForSEO przyjmuje TYLKO JEDNO zadanie
    na request ("You can set only one task at a time.") — więc każda fraza to osobne
    zapytanie HTTP, wysyłane pojedynczo, a nie w paczkach.

    Każda fraza jest też pytana OSOBNO od pozostałych (bez porównywania kilku fraz naraz),
    bo Google Trends przy porównaniu normalizuje wynik względem najsilniejszej frazy w grupie
    i potrafi zwrócić mnóstwo braków danych dla słabszych fraz — psuje to wykres. Pojedyncze
    zapytania dają dokładnie taki wykres, jak ręczne sprawdzenie frazy na trends.google.com.

    Domyślnie pobiera 5 lat historii (jak w trends.google.com). Dłuższy timeout i ponowne
    próby na wypadek wolnej odpowiedzi; jedno niepowodzenie nie przerywa całego biegu."""
    headers = dfs_auth_header(login, password)
    result: dict[str, pd.DataFrame] = {}
    failed: list[str] = []
    reasons: dict[str, str] = {}  # fraza -> prawdziwy powód niepowodzenia (diagnostyka)

    for i, kw in enumerate(keywords):
        if progress_callback:
            progress_callback(i, len(keywords), kw)

        payload = [{
            "keywords": [kw],
            "location_code": location_code,
            "time_range": time_range,
            "type": "web",
        }]

        resp = None
        last_exc = None
        for attempt in range(max_retries + 1):
            try:
                resp = requests.post(f"{DFS_BASE}/keywords_data/google_trends/explore/live",
                                      headers=headers, json=payload, timeout=60)
                resp.raise_for_status()
                break
            except requests.exceptions.RequestException as e:
                last_exc = e
                resp = None
                if attempt < max_retries:
                    time.sleep(2)
                continue

        if resp is None:
            detail = str(last_exc)
            if isinstance(last_exc, requests.exceptions.HTTPError) and last_exc.response is not None:
                detail = f"HTTP {last_exc.response.status_code}: {last_exc.response.text[:200]}"
            reasons[kw] = detail
            failed.append(kw)
            continue

        data = resp.json()
        tasks = data.get("tasks", [])
        found = False
        task_status = "brak wyniku w odpowiedzi API"
        for task in tasks:
            task_status = f"{task.get('status_code')}: {task.get('status_message')}"
            for item in (task.get("result") or []):
                if item is None:
                    continue
                # rzeczywista struktura DataForSEO:
                # result[] -> items[] -> {"type": "google_trends_graph", "keywords": [...], "data": [{"date_from":..,"values":[..]}]}
                for sub in (item.get("items") or []):
                    if sub.get("type") != "google_trends_graph":
                        continue
                    kws = sub.get("keywords") or [kw]
                    series_data = sub.get("data") or []
                    dates, values = [], []
                    for point in series_data:
                        try:
                            v = point.get("values", [None] * len(kws))[0]
                        except (IndexError, TypeError):
                            v = None
                        dates.append(point.get("date_from"))
                        values.append(v)
                    if dates:
                        tdf = pd.DataFrame({"data": pd.to_datetime(dates, errors="coerce"), "trend": values})
                        tdf = tdf.dropna(subset=["data"]).sort_values("data").reset_index(drop=True)
                        result[kw] = tdf
                        found = True

        if not found:
            failed.append(kw)
            reasons[kw] = task_status
        time.sleep(0.15)  # uprzejmie dla rate limitu

    return result, failed, reasons


# ---------------------------------------------------------------------------
# Senuto (best-effort — dostosuj do swojego planu API)
# ---------------------------------------------------------------------------

def senuto_get_token(email: str, password: str) -> str:
    resp = requests.post(f"{SENUTO_BASE}/users/token", json={"email": email, "password": password}, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    # nazwa pola tokena może się różnić w zależności od wersji API — sprawdź docs-api.senuto.com
    return data.get("token") or data.get("access_token") or data.get("bearer_token")


def senuto_get_volumes(bearer_token: str, keywords: list[str]) -> pd.DataFrame:
    """SZKIELET integracji. Endpoint/parametry do potwierdzenia w Twojej dokumentacji Senuto API
    (Postman collection na docs-api.senuto.com, sekcja Keyword Database)."""
    headers = {"Authorization": f"Bearer {bearer_token}", "Content-Type": "application/json"}
    rows = []
    for kw in keywords:
        try:
            resp = requests.get(f"{SENUTO_BASE}/keyword_database/keywords",
                                 headers=headers, params={"keyword": kw}, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            item = (data.get("data") or data.get("items") or [None])[0]
            rows.append({
                "fraza": kw,
                "senuto_wolumen": item.get("volume") if item else None,
            })
        except Exception:
            rows.append({"fraza": kw, "senuto_wolumen": None})
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Sygnał z Google Trends (osobno od wolumenu Google Ads)
# ---------------------------------------------------------------------------

def compute_trend_signal(trends: dict[str, pd.DataFrame], threshold_pct: float = 10.0) -> pd.DataFrame:
    """Dla każdej frazy porównuje średnią popularność w pierwszej i drugiej połowie
    pobranego okresu Google Trends — daje niezależne od Google Ads potwierdzenie,
    czy zainteresowanie frazą rośnie czy spada."""
    rows = []
    for phrase, tdf in trends.items():
        if tdf is None or tdf.empty or len(tdf) < 4:
            continue
        d = tdf.dropna(subset=["trend"])
        if len(d) < 4:
            continue
        mid = len(d) // 2
        early = d["trend"].iloc[:mid].mean()
        late = d["trend"].iloc[mid:].mean()
        change = round((late - early) / early * 100, 1) if early else None
        if change is None:
            kierunek = "brak danych"
        elif change > threshold_pct:
            kierunek = "rosnący"
        elif change < -threshold_pct:
            kierunek = "malejący"
        else:
            kierunek = "stabilny"
        rows.append({
            "fraza": phrase,
            "trend_google_wczesniej": round(early, 1),
            "trend_google_ostatnio": round(late, 1),
            "trend_google_zmiana_%": change,
            "trend_google_kierunek": kierunek,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Podsumowanie dla klienta
# ---------------------------------------------------------------------------

def generate_summary(df: pd.DataFrame, threshold_pct: float = 5.0) -> dict | None:
    """Zestawia wynik na poziomie całej listy fraz: czy trend wyszukiwań w tym roku
    jest wyższy, niższy czy stabilny względem roku poprzedniego."""
    d = df.dropna(subset=["wolumen_aktualny", "wolumen_rok_temu", "zmiana_%"]).copy()
    if d.empty:
        return None

    total_now = d["wolumen_aktualny"].sum()
    total_before = d["wolumen_rok_temu"].sum()
    overall_change = round((total_now - total_before) / total_before * 100, 1) if total_before else None

    up = d[d["zmiana_%"] > threshold_pct]
    down = d[d["zmiana_%"] < -threshold_pct]
    stable = d[(d["zmiana_%"] >= -threshold_pct) & (d["zmiana_%"] <= threshold_pct)]

    if overall_change is None:
        kierunek = "brak danych"
    elif overall_change > threshold_pct:
        kierunek = "wyższy"
    elif overall_change < -threshold_pct:
        kierunek = "niższy"
    else:
        kierunek = "na podobnym poziomie"

    out = {
        "liczba_fraz": len(d),
        "total_now": int(total_now),
        "total_before": int(total_before),
        "overall_change": overall_change,
        "kierunek": kierunek,
        "liczba_rosnacych": len(up),
        "liczba_spadajacych": len(down),
        "liczba_stabilnych": len(stable),
        "top_rosnace": d.sort_values("zmiana_%", ascending=False).head(5)[["fraza", "zmiana_%"]],
        "top_spadajace": d.sort_values("zmiana_%", ascending=True).head(5)[["fraza", "zmiana_%"]],
        "all_rosnace": up.sort_values("zmiana_%", ascending=False)[["fraza", "zmiana_%"]],
        "all_spadajace": down.sort_values("zmiana_%", ascending=True)[["fraza", "zmiana_%"]],
    }

    # niezależne potwierdzenie z Google Trends (jeśli dane są dostępne w df)
    if "trend_google_kierunek" in df.columns:
        tg = df.dropna(subset=["trend_google_kierunek"])
        if not tg.empty:
            out["trend_google_rosnacych"] = int((tg["trend_google_kierunek"] == "rosnący").sum())
            out["trend_google_malejacych"] = int((tg["trend_google_kierunek"] == "malejący").sum())
            out["trend_google_stabilnych"] = int((tg["trend_google_kierunek"] == "stabilny").sum())

    return out


# ---------------------------------------------------------------------------
# Analiza konkurencji (frazy brandowe przypisane do konkretnych konkurentów)
# ---------------------------------------------------------------------------

def compute_competitor_summary(competitor_df: pd.DataFrame) -> pd.DataFrame:
    """Grupuje wolumen wyszukiwań fraz brandowych po konkurencie."""
    if competitor_df is None or competitor_df.empty:
        return pd.DataFrame()
    d = competitor_df.dropna(subset=["wolumen_aktualny"]).copy()
    if d.empty:
        return pd.DataFrame()
    grouped = d.groupby("konkurent", as_index=False).agg(
        liczba_fraz=("fraza", "count"),
        wolumen_teraz=("wolumen_aktualny", "sum"),
        wolumen_rok_temu=("wolumen_rok_temu", "sum"),
    )
    grouped["zmiana_%"] = grouped.apply(
        lambda r: round((r["wolumen_teraz"] - r["wolumen_rok_temu"]) / r["wolumen_rok_temu"] * 100, 1)
        if r["wolumen_rok_temu"] else None, axis=1,
    )
    return grouped.sort_values("wolumen_teraz", ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Eksport: Excel (podsumowanie + tabela + wykres) i PDF (raport dla klienta)
# ---------------------------------------------------------------------------

def _dark_fig(figsize):
    """Figura matplotlib w kolorystyce Odyseo: czarne tło, biały tekst."""
    fig, ax = plt.subplots(figsize=figsize)
    fig.patch.set_facecolor("black")
    ax.set_facecolor("black")
    for spine in ax.spines.values():
        spine.set_color("#555555")
    ax.tick_params(colors="white", labelsize=8)
    ax.title.set_color("white")
    ax.xaxis.label.set_color("white")
    ax.yaxis.label.set_color("white")
    return fig, ax


def _total_volume_chart_png(summary: dict) -> io.BytesIO:
    fig, ax = _dark_fig((5, 3))
    labels = ["Rok temu", "Teraz"]
    values = [summary["total_before"], summary["total_now"]]
    ax.bar(labels, values, color=["#9CA3AF", ODYSEO_PURPLE_LIGHT])
    ax.set_ylabel("Łączny wolumen wyszukiwań / mies.")
    ax.set_title("Łączny wolumen — porównanie rok do roku")
    for i, v in enumerate(values):
        ax.text(i, v, f"{v:,}".replace(",", " "), ha="center", va="bottom", fontsize=9, color="white")
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, facecolor="black")
    plt.close(fig)
    buf.seek(0)
    return buf


def _movers_chart_png(up_df: pd.DataFrame, down_df: pd.DataFrame, title: str) -> io.BytesIO | None:
    combo = pd.concat([up_df, down_df]).drop_duplicates(subset="fraza")
    if combo.empty:
        return None
    combo = combo.sort_values("zmiana_%")
    fig, ax = _dark_fig((7, max(3, 0.35 * len(combo))))
    bar_colors = ["#EF4444" if v < 0 else "#22C55E" for v in combo["zmiana_%"]]
    ax.barh(combo["fraza"], combo["zmiana_%"], color=bar_colors)
    ax.axvline(0, color="white", linewidth=0.8)
    ax.set_xlabel("Zmiana r/r (%)")
    ax.set_title(title)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, facecolor="black")
    plt.close(fig)
    buf.seek(0)
    return buf


def _trend_line_chart_png(phrase: str, tdf: pd.DataFrame, figsize=(5.5, 2.3)) -> io.BytesIO | None:
    if tdf is None or tdf.empty:
        return None
    d = tdf.copy()
    d["data"] = pd.to_datetime(d["data"], errors="coerce")
    d = d.dropna(subset=["data"]).sort_values("data")
    if d.empty or d["trend"].notna().sum() < 4:
        # Google Trends ma za mało realnych punktów dla tej frazy (prawie same braki/missing_data,
        # zostaje 0-3 punkty) — linii i tak nie da się sensownie narysować, więc pomijamy wykres
        # zamiast rysować mylącą, pustą ramkę z przypadkowo dobraną skalą osi.
        return None
    fig, ax = _dark_fig(figsize)
    ax.plot(d["data"], d["trend"], color=ODYSEO_PURPLE_LIGHT, linewidth=1.5)
    ax.set_title(phrase, fontsize=9)
    fig.autofmt_xdate()
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, facecolor="black")
    plt.close(fig)
    buf.seek(0)
    return buf


def _competitor_chart_png(competitor_summary: pd.DataFrame) -> io.BytesIO | None:
    if competitor_summary is None or competitor_summary.empty:
        return None
    d = competitor_summary.sort_values("wolumen_teraz")
    fig, ax = _dark_fig((7, max(3, 0.5 * len(d))))
    ax.barh(d["konkurent"], d["wolumen_teraz"], color=ODYSEO_PURPLE_LIGHT)
    ax.set_xlabel("Łączny wolumen wyszukiwań fraz brandowych (teraz) / mies.")
    ax.set_title("Analiza konkurencji — wolumen fraz brandowych")
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, facecolor="black")
    plt.close(fig)
    buf.seek(0)
    return buf


def generate_excel_report(df: pd.DataFrame, summary: dict | None, client_name: str = "",
                           competitor_df: pd.DataFrame | None = None,
                           competitor_summary: pd.DataFrame | None = None) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dane szczegółowe")

        if competitor_df is not None and not competitor_df.empty:
            competitor_df.to_excel(writer, index=False, sheet_name="Konkurencja — frazy")
            if competitor_summary is not None and not competitor_summary.empty:
                competitor_summary.to_excel(writer, index=False, sheet_name="Konkurencja — podsumowanie")

        if summary is not None:
            ws = writer.book.create_sheet("Podsumowanie", 0)
            bold = Font(bold=True, color="422AAB")
            title = Font(bold=True, size=14, color="422AAB")

            title_text = "Trendomat by Odyseo — podsumowanie fraz"
            if client_name:
                title_text += f" ({client_name})"
            ws["A1"] = title_text
            ws["A1"].font = title
            ws["A2"] = f"Wygenerowano: {dt.date.today().isoformat()}"

            ws["A4"] = "Sprawdzonych fraz"
            ws["B4"] = summary["liczba_fraz"]
            ws["A5"] = "Wolumen rok temu (łącznie/mies.)"
            ws["B5"] = summary["total_before"]
            ws["A6"] = "Wolumen teraz (łącznie/mies.)"
            ws["B6"] = summary["total_now"]
            ws["A7"] = "Zmiana r/r"
            ws["B7"] = f"{summary['overall_change']:+.1f}%"
            ws["A8"] = "Kierunek"
            ws["B8"] = summary["kierunek"]
            for r in range(4, 9):
                ws[f"A{r}"].font = bold

            ws["A10"] = "Frazy rosnące"
            ws["B10"] = summary["liczba_rosnacych"]
            ws["A11"] = "Frazy spadające"
            ws["B11"] = summary["liczba_spadajacych"]
            ws["A12"] = "Frazy stabilne"
            ws["B12"] = summary["liczba_stabilnych"]
            for r in range(10, 13):
                ws[f"A{r}"].font = bold

            ws["D4"] = "Łączny wolumen"
            ws["D4"].font = bold
            ws["D5"] = "Rok temu"
            ws["E5"] = summary["total_before"]
            ws["D6"] = "Teraz"
            ws["E6"] = summary["total_now"]

            chart = BarChart()
            chart.title = "Łączny wolumen — rok do roku"
            chart.y_axis.title = "Wyszukiwania / mies."
            data_ref = Reference(ws, min_col=5, min_row=5, max_row=6)
            cats_ref = Reference(ws, min_col=4, min_row=5, max_row=6)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = "422AAB"
            chart.width = 12
            chart.height = 7
            ws.add_chart(chart, "D9")

            for col, width in zip("ABCDE", (32, 16, 4, 12, 12)):
                ws.column_dimensions[col].width = width

    return buf.getvalue()


def _fmt_num(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    return f"{int(v):,}".replace(",", " ")


def _fmt_pct(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    return f"{v:+.1f}%"


def _fmt_str(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    return str(v)


def _pdf_bg(canvas, doc_):
    """Czarne tło na każdej stronie PDF."""
    canvas.saveState()
    canvas.setFillColor(PDF_BLACK)
    canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
    canvas.restoreState()


def generate_pdf_report(df: pd.DataFrame, summary: dict | None, trends: dict, client_name: str = "",
                         competitor_df: pd.DataFrame | None = None,
                         competitor_summary: pd.DataFrame | None = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.3 * cm, rightMargin=1.3 * cm)

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontName="DejaVuSans-Bold",
                         fontSize=19, spaceAfter=6, textColor=PDF_WHITE)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName="DejaVuSans-Bold",
                         fontSize=14, spaceBefore=14, spaceAfter=8, textColor=PDF_PURPLE_LIGHT)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontName="DejaVuSans",
                           fontSize=10, textColor=PDF_WHITE, leading=14)
    footer_style = ParagraphStyle("footer", parent=body, fontSize=8, textColor=colors.HexColor("#9CA3AF"))
    answer_style = ParagraphStyle("answer", parent=body, fontName="DejaVuSans-Bold",
                                   fontSize=16, textColor=PDF_WHITE, alignment=1, leading=22)
    th_style = ParagraphStyle("th", parent=body, fontName="DejaVuSans-Bold", fontSize=8.5,
                               textColor=PDF_WHITE, leading=10)

    def table_style(header_bg=PDF_PURPLE, header_text=PDF_WHITE):
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), header_text),
            ("TEXTCOLOR", (0, 1), (-1, -1), PDF_WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, PDF_GRID),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PDF_BLACK, colors.HexColor("#0D0D0D")]),
        ])

    title_text = "Trendomat by Odyseo — raport trendów wyszukiwań"
    story = [
        Paragraph(title_text, h1),
    ]
    if client_name:
        story.append(Paragraph(f"Klient / marka: {client_name}", footer_style))
    story.append(Paragraph(f"Wygenerowano: {dt.date.today().strftime('%d.%m.%Y')}", footer_style))
    story.append(Spacer(1, 0.5 * cm))

    # ------------------------------------------------------------------
    # METODOLOGIA — skąd pochodzą dane i jak działają narzędzia
    # ------------------------------------------------------------------
    story.append(Paragraph("Skąd pochodzą dane w tym raporcie", h2))
    story.append(Paragraph(
        "Dane pobierane są automatycznie z API <b>DataForSEO</b> — zewnętrznego dostawcy danych "
        "wyszukiwania, który agreguje i udostępnia w formie API te same źródła, z których korzystają "
        "narzędzia takie jak Google Keyword Planner czy Google Trends.", body))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "<b>Wolumen wyszukiwań</b> (kolumny „Wolumen teraz” i „Wolumen rok temu”) pochodzi z bazy "
        "<b>Google Ads</b> (ten sam system danych co Google Keyword Planner) — to średnia miesięczna "
        "liczba wyszukiwań danej frazy w Google, szacowana przez samego Google na podstawie "
        "rzeczywistego ruchu w wyszukiwarce. „Wolumen rok temu” wyliczany jest z 12-miesięcznej "
        "historii wyszukiwań, jaką Google Ads udostępnia dla każdej frazy.", body))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "<b>Kolumna „Trend Google”</b> oraz wykresy w rozdziale 1 pochodzą z <b>Google Trends</b> — "
        "to nie jest liczba wyszukiwań, tylko względna popularność frazy w czasie w skali 0–100 "
        "(100 = szczyt popularności tej frazy w badanym okresie). Każda fraza jest sprawdzana "
        "osobno, dokładnie tak jak przy ręcznym wyszukaniu pojedynczej frazy na trends.google.com. "
        "Dla części rzadko wyszukiwanych fraz Google Trends nie ma wystarczających danych, żeby "
        "wyznaczyć wiarygodny trend — takie przypadki są w raporcie oznaczone jako „-” lub pomijane "
        "na wykresach (zamiast pokazywać mylące, puste wykresy).", body))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "Kierunek trendu („rosnący”/„malejący”/„stabilny”) liczony jest przez porównanie średniej "
        "popularności z pierwszej i drugiej połowy badanego okresu (domyślnie 5 lat wstecz, tak jak "
        "w standardowym widoku trends.google.com).", body))
    story.append(Spacer(1, 0.4 * cm))

    # ------------------------------------------------------------------
    # ROZDZIAŁ 1: zebrane dane — pełna tabela + wszystkie wykresy Trends
    # ------------------------------------------------------------------
    story.append(Paragraph("1. Zebrane dane", h1))
    story.append(Paragraph(
        f"Pełne zestawienie dla wszystkich {len(df)} sprawdzonych fraz: wolumen wyszukiwań "
        f"aktualny i sprzed roku oraz sygnał z Google Trends.", body))
    story.append(Spacer(1, 0.3 * cm))

    has_trend_col = "trend_google_kierunek" in df.columns
    header_labels = ["Fraza", "Wolumen teraz", "Wolumen rok temu", "Zmiana r/r"]
    if has_trend_col:
        header_labels.append("Trend Google")
    header_row = [Paragraph(h, th_style) for h in header_labels]
    data_rows = [header_row]
    for _, r in df.iterrows():
        row = [r.get("fraza", ""), _fmt_num(r.get("wolumen_aktualny")),
               _fmt_num(r.get("wolumen_rok_temu")), _fmt_pct(r.get("zmiana_%"))]
        if has_trend_col:
            row.append(_fmt_str(r.get("trend_google_kierunek")))
        data_rows.append(row)
    col_widths = [7.2 * cm, 2.6 * cm, 2.9 * cm, 2.4 * cm] + ([2.2 * cm] if has_trend_col else [])
    full_table = Table(data_rows, colWidths=col_widths, repeatRows=1)
    full_table.setStyle(table_style())
    story.append(full_table)
    story.append(PageBreak())

    # renderujemy wykresy tylko dla fraz, dla których Google Trends faktycznie zwrócił dane
    # (dla części długiego ogona zapytań Google zwraca same braki — pusty wykres nic by nie wnosił)
    trend_charts = []
    for phrase in df["fraza"]:
        png = _trend_line_chart_png(phrase, trends.get(phrase))
        if png is not None:
            trend_charts.append((phrase, png))

    no_data_count = len(df) - len(trend_charts)
    if trend_charts:
        story.append(Paragraph(
            f"Wykresy Google Trends — frazy z dostępnymi danymi ({len(trend_charts)} z {len(df)})", h2))
        if no_data_count:
            story.append(Paragraph(
                f"Dla {no_data_count} fraz Google Trends nie zwrócił żadnych danych (zbyt niski/rzadki "
                f"wolumen wyszukiwań, żeby oszacować popularność w czasie) — pominięto puste wykresy.", body))
        story.append(Spacer(1, 0.2 * cm))
        img_w, img_h = 8.3 * cm, 3.6 * cm
        for j in range(0, len(trend_charts), 2):
            pair = trend_charts[j:j + 2]
            imgs = [Image(png, width=img_w, height=img_h) for _, png in pair]
            if len(imgs) == 1:
                imgs.append(Paragraph("", body))
            row_table = Table([imgs], colWidths=[img_w + 0.3 * cm, img_w + 0.3 * cm])
            row_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            story.append(row_table)
        story.append(PageBreak())
    else:
        story.append(Paragraph("Brak dostępnych wykresów Google Trends dla sprawdzonych fraz.", body))
        story.append(PageBreak())

    # ------------------------------------------------------------------
    # ROZDZIAŁ 2: trendy — frazy rosnące i spadające
    # ------------------------------------------------------------------
    story.append(Paragraph("2. Trendy: frazy rosnące i spadające", h1))
    story.append(Spacer(1, 0.2 * cm))

    if summary is None:
        story.append(Paragraph("Za mało danych, żeby wyznaczyć rosnące i spadające frazy.", body))
    else:
        all_up = summary["all_rosnace"]
        all_down = summary["all_spadajace"]

        movers_buf = _movers_chart_png(all_up, all_down, "Frazy rosnące i spadające — zmiana r/r")
        if movers_buf is not None:
            story.append(Image(movers_buf, width=15 * cm,
                                height=15 * cm * 0.6 if len(all_up) + len(all_down) < 15 else 20 * cm))
            story.append(Spacer(1, 0.4 * cm))

        story.append(Paragraph(f"Frazy rosnące ({len(all_up)})", h2))
        if all_up.empty:
            story.append(Paragraph("Brak fraz rosnących powyżej progu istotności.", body))
        else:
            up_rows = [["Fraza", "Zmiana r/r"]] + [
                [r["fraza"], _fmt_pct(r["zmiana_%"])] for _, r in all_up.iterrows()
            ]
            t_up = Table(up_rows, colWidths=[11 * cm, 4 * cm], repeatRows=1)
            t_up.setStyle(table_style(header_bg=colors.HexColor("#14532D")))
            story.append(t_up)
        story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph(f"Frazy spadające ({len(all_down)})", h2))
        if all_down.empty:
            story.append(Paragraph("Brak fraz spadających powyżej progu istotności.", body))
        else:
            down_rows = [["Fraza", "Zmiana r/r"]] + [
                [r["fraza"], _fmt_pct(r["zmiana_%"])] for _, r in all_down.iterrows()
            ]
            t_down = Table(down_rows, colWidths=[11 * cm, 4 * cm], repeatRows=1)
            t_down.setStyle(table_style(header_bg=colors.HexColor("#7F1D1D")))
            story.append(t_down)

    story.append(PageBreak())

    # ------------------------------------------------------------------
    # ROZDZIAŁ 3: podsumowanie — najpierw odpowiedź, potem uzasadnienie
    # ------------------------------------------------------------------
    story.append(Paragraph("3. Podsumowanie", h1))
    story.append(Spacer(1, 0.3 * cm))

    if summary is None:
        story.append(Paragraph("Za mało danych, żeby policzyć podsumowanie.", body))
    else:
        odpowiedz_map = {
            "wyższy": "WYŻSZY",
            "niższy": "NIŻSZY",
            "na podobnym poziomie": "NA PODOBNYM POZIOMIE",
            "brak danych": "BRAK DANYCH",
        }
        odpowiedz = odpowiedz_map.get(summary["kierunek"], summary["kierunek"].upper())

        answer_table = Table(
            [[Paragraph("Czy dla wybranych słów kluczowych trend wyszukiwań w tym roku "
                        "jest mniejszy, czy większy?", ParagraphStyle(
                            "q", parent=body, alignment=1, fontName="DejaVuSans-Bold", fontSize=11))],
             [Paragraph(odpowiedz, answer_style)]],
            colWidths=[16 * cm],
        )
        answer_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), PDF_PURPLE),
            ("BOX", (0, 0), (-1, -1), 1, PDF_PURPLE_LIGHT),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(answer_table)
        story.append(Spacer(1, 0.6 * cm))

        story.append(Paragraph("Uzasadnienie", h2))
        zdanie = (
            f"Dla sprawdzonych {summary['liczba_fraz']} fraz łączny wolumen wyszukiwań jest "
            f"<b>{summary['kierunek']}</b> niż rok temu "
            f"({summary['overall_change']:+.1f}%, "
            f"{_fmt_num(summary['total_before'])} → {_fmt_num(summary['total_now'])} "
            f"wyszukiwań miesięcznie łącznie)."
        )
        story.append(Paragraph(zdanie, body))
        story.append(Spacer(1, 0.3 * cm))

        metrics_rows = [["Frazy rosnące", "Frazy spadające", "Frazy stabilne"],
                         [str(summary["liczba_rosnacych"]), str(summary["liczba_spadajacych"]),
                          str(summary["liczba_stabilnych"])]]
        metrics_table = Table(metrics_rows, colWidths=[5.3 * cm] * 3)
        metrics_table.setStyle(table_style())
        story.append(metrics_table)
        story.append(Spacer(1, 0.3 * cm))

        if "trend_google_rosnacych" in summary:
            story.append(Paragraph(
                "Niezależne potwierdzenie z Google Trends (popularność wyszukiwań w czasie):", body))
            gt_rows = [["Rosnący trend Google", "Malejący trend Google", "Stabilny trend Google"],
                       [str(summary["trend_google_rosnacych"]), str(summary["trend_google_malejacych"]),
                        str(summary["trend_google_stabilnych"])]]
            gt_table = Table(gt_rows, colWidths=[5.3 * cm] * 3)
            gt_table.setStyle(table_style())
            story.append(gt_table)
            story.append(Spacer(1, 0.4 * cm))

        chart_buf = _total_volume_chart_png(summary)
        story.append(Image(chart_buf, width=12 * cm, height=7.2 * cm))

    # ------------------------------------------------------------------
    # ROZDZIAŁ 4: analiza konkurencji (opcjonalny) — tylko wolumen z DataForSEO,
    # bez Google Trends, frazy pogrupowane wg przypisanego konkurenta
    # ------------------------------------------------------------------
    if competitor_df is not None and not competitor_df.empty:
        story.append(PageBreak())
        story.append(Paragraph("4. Analiza marek konkurencji", h1))
        story.append(Paragraph(
            "Wolumen wyszukiwań fraz brandowych przypisanych do poszczególnych konkurentów "
            "(źródło: DataForSEO / Google Ads — bez Google Trends).", body))
        story.append(Spacer(1, 0.3 * cm))

        chart_buf = _competitor_chart_png(competitor_summary)
        if chart_buf is not None:
            story.append(Image(chart_buf, width=15 * cm,
                                height=15 * cm * min(1.0, 0.15 * max(3, len(competitor_summary)))))
            story.append(Spacer(1, 0.4 * cm))

        def hdr(*labels):
            return [Paragraph(lbl, th_style) for lbl in labels]

        if competitor_summary is not None and not competitor_summary.empty:
            story.append(Paragraph("Podsumowanie wg konkurenta", h2))
            sum_rows = [hdr("Konkurent", "Liczba fraz", "Wolumen teraz", "Wolumen rok temu", "Zmiana r/r")] + [
                [r["konkurent"], str(int(r["liczba_fraz"])), _fmt_num(r["wolumen_teraz"]),
                 _fmt_num(r["wolumen_rok_temu"]), _fmt_pct(r["zmiana_%"])]
                for _, r in competitor_summary.iterrows()
            ]
            sum_table = Table(sum_rows, colWidths=[4.5 * cm, 2.6 * cm, 3.1 * cm, 3.3 * cm, 2.5 * cm], repeatRows=1)
            sum_table.setStyle(table_style())
            story.append(sum_table)
            story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("Szczegóły — frazy wg konkurenta", h2))
        det_rows = [hdr("Konkurent", "Fraza", "Wolumen teraz", "Wolumen rok temu", "Zmiana r/r")]
        for _, r in competitor_df.sort_values(["konkurent", "fraza"]).iterrows():
            det_rows.append([
                r.get("konkurent", ""), r.get("fraza", ""),
                _fmt_num(r.get("wolumen_aktualny")), _fmt_num(r.get("wolumen_rok_temu")),
                _fmt_pct(r.get("zmiana_%")),
            ])
        det_table = Table(det_rows, colWidths=[3.2 * cm, 5.6 * cm, 2.7 * cm, 3 * cm, 2.5 * cm], repeatRows=1)
        det_table.setStyle(table_style())
        story.append(det_table)

    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph(
        "Źródło danych: DataForSEO (Google Ads — wolumen wyszukiwań, Google Trends — popularność 0–100). "
        "Wolumen „rok temu” pochodzi z historii miesięcznej ostatnich 12 miesięcy.",
        footer_style,
    ))

    doc.build(story, onFirstPage=_pdf_bg, onLaterPages=_pdf_bg)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.title("🧭 Trendomat by Odyseo")
st.caption("Wolumen wyszukiwań (aktualny i sprzed roku) oraz trend Google dla dowolnej listy fraz — "
           "dla dowolnego klienta i marki.")

with st.sidebar:
    st.header("Klucze API")

    st.subheader("DataForSEO")
    dfs_login = st.text_input("DataForSEO login (e-mail)", type="default")
    dfs_password = st.text_input("DataForSEO password", type="password")

    st.subheader("Senuto (opcjonalnie)")
    use_senuto = st.checkbox("Dołącz dane z Senuto", value=False)
    senuto_email = st.text_input("Senuto e-mail", disabled=not use_senuto)
    senuto_password = st.text_input("Senuto hasło", type="password", disabled=not use_senuto)

    st.subheader("Ustawienia")
    client_name = st.text_input("Nazwa klienta / marki (opcjonalnie)",
                                 placeholder="np. Aura Herbals", help="Pojawi się w tytule raportu PDF/Excel.")
    location_code = st.number_input("Kod lokalizacji DataForSEO (Polska = 2616)", value=2616, step=1)
    language_code = st.text_input("Kod języka DataForSEO", value="pl")
    trends_range = st.selectbox(
        "Okres Google Trends",
        options=["past_5_years", "past_12_months", "past_90_days"],
        format_func=lambda v: {"past_5_years": "5 lat (jak w trends.google.com)",
                                "past_12_months": "12 miesięcy",
                                "past_90_days": "90 dni"}[v],
        index=0,
    )

st.subheader("1. Wklej listę fraz (jedna fraza w linii, max 50)")
raw_phrases = st.text_area("Frazy", height=220, placeholder="fraza kluczowa 1\nfraza kluczowa 2\n...")

phrases = parse_phrases(raw_phrases)
if phrases:
    st.caption(f"Wykryto {len(phrases)} unikalnych fraz.")
    if len(phrases) > 50:
        st.warning("Wykryto więcej niż 50 fraz — zostaną przetworzone wszystkie, ale sprawdź listę.")

run = st.button("2. Pobierz dane", type="primary", disabled=not phrases)

if "report_df" not in st.session_state:
    st.session_state.report_df = None
if "trends_data" not in st.session_state:
    st.session_state.trends_data = {}

if run:
    if not dfs_login or not dfs_password:
        st.error("Podaj login i hasło do DataForSEO w panelu bocznym.")
    else:
        with st.spinner("Pobieram wolumen wyszukiwań (DataForSEO)..."):
            try:
                vol_df = dfs_search_volume(dfs_login, dfs_password, phrases, int(location_code), language_code)
            except Exception as e:
                st.error(f"Błąd DataForSEO (search_volume): {e}")
                vol_df = pd.DataFrame()

        st.caption("Pobieram Google Trends (DataForSEO) — każda fraza to osobne zapytanie, może to potrwać kilka minut.")
        trends_progress = st.progress(0.0)
        trends_status = st.empty()

        def _trends_progress(i, total, kw):
            trends_progress.progress((i + 1) / total)
            trends_status.text(f"{i + 1}/{total}: {kw}")

        try:
            trends, failed_trends, trend_reasons = dfs_trends_explore(
                dfs_login, dfs_password, phrases, int(location_code), time_range=trends_range,
                progress_callback=_trends_progress)
            trends_progress.empty()
            trends_status.empty()
            if failed_trends:
                unique_reasons = sorted(set(trend_reasons.get(kw, "nieznany powód") for kw in failed_trends))
                st.warning(
                    f"Nie udało się pobrać Google Trends dla {len(failed_trends)} fraz: "
                    f"{', '.join(failed_trends)}.\n\n"
                    f"Powód(y) zwrócone przez API: {' | '.join(unique_reasons[:5])}"
                    + (" (i inne)" if len(unique_reasons) > 5 else "") + ". "
                    f"Reszta danych jest kompletna — spróbuj ponownie, jeśli te frazy są kluczowe."
                )
        except Exception as e:
            trends_progress.empty()
            trends_status.empty()
            st.error(f"Błąd DataForSEO (google_trends/explore): {e}")
            trends = {}

        senuto_df = pd.DataFrame()
        if use_senuto:
            if not senuto_email or not senuto_password:
                st.warning("Zaznaczono Senuto, ale brak danych logowania — pomijam ten krok.")
            else:
                with st.spinner("Pobieram dane z Senuto..."):
                    try:
                        token = senuto_get_token(senuto_email, senuto_password)
                        senuto_df = senuto_get_volumes(token, phrases)
                    except Exception as e:
                        st.warning(f"Nie udało się pobrać danych z Senuto: {e}. "
                                   f"Sprawdź endpoint w swojej dokumentacji API (docs-api.senuto.com).")

        merged = vol_df
        if not senuto_df.empty:
            merged = merged.merge(senuto_df, on="fraza", how="left")
        trend_signal_df = compute_trend_signal(trends)
        if not trend_signal_df.empty:
            merged = merged.merge(trend_signal_df, on="fraza", how="left")

        st.session_state.report_df = merged
        st.session_state.trends_data = trends

# ---------------------------------------------------------------------------
# Wyniki
# ---------------------------------------------------------------------------

if st.session_state.report_df is not None and not st.session_state.report_df.empty:
    df = st.session_state.report_df

    st.subheader("3. Podsumowanie dla klienta")
    summary = generate_summary(df)
    if summary is None:
        st.info("Za mało danych (wolumen aktualny / sprzed roku), żeby policzyć podsumowanie.")
    else:
        box = st.success if summary["kierunek"] == "wyższy" else (
            st.error if summary["kierunek"] == "niższy" else st.info)
        zdanie = (
            f"Dla sprawdzonych {summary['liczba_fraz']} fraz łączny wolumen wyszukiwań jest "
            f"**{summary['kierunek']}** niż rok temu "
            f"({summary['overall_change']:+.1f}%, {summary['total_before']:,} → {summary['total_now']:,} "
            f"wyszukiwań miesięcznie łącznie)."
        ).replace(",", " ")
        box(zdanie)

        m1, m2, m3 = st.columns(3)
        m1.metric("Frazy rosnące (wolumen)", summary["liczba_rosnacych"])
        m2.metric("Frazy spadające (wolumen)", summary["liczba_spadajacych"])
        m3.metric("Frazy stabilne (wolumen)", summary["liczba_stabilnych"])

        if "trend_google_rosnacych" in summary:
            st.caption("Niezależne potwierdzenie z Google Trends (popularność wyszukiwań w czasie):")
            g1, g2, g3 = st.columns(3)
            g1.metric("Rosnący trend Google", summary["trend_google_rosnacych"])
            g2.metric("Malejący trend Google", summary["trend_google_malejacych"])
            g3.metric("Stabilny trend Google", summary["trend_google_stabilnych"])

        c1, c2 = st.columns(2)
        with c1:
            st.caption("Najmocniej rosnące frazy")
            st.dataframe(summary["top_rosnace"], hide_index=True, use_container_width=True)
        with c2:
            st.caption("Najmocniej spadające frazy")
            st.dataframe(summary["top_spadajace"], hide_index=True, use_container_width=True)

    st.subheader("4. Szczegółowa tabela")
    st.dataframe(df, use_container_width=True, hide_index=True)

    file_slug = slugify(client_name) if client_name else "trendomat"
    comp_df_for_export = st.session_state.get("competitor_df")
    comp_summary_for_export = st.session_state.get("competitor_summary")

    col1, col2, col3 = st.columns(3)
    with col1:
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("Pobierz CSV", data=csv_bytes, file_name=f"{file_slug}_raport_fraz.csv", mime="text/csv")
    with col2:
        excel_bytes = generate_excel_report(df, summary, client_name=client_name,
                                             competitor_df=comp_df_for_export,
                                             competitor_summary=comp_summary_for_export)
        st.download_button("Pobierz Excel (z podsumowaniem i wykresem)", data=excel_bytes,
                            file_name=f"{file_slug}_raport_fraz.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col3:
        pdf_bytes = generate_pdf_report(df, summary, st.session_state.trends_data, client_name=client_name,
                                         competitor_df=comp_df_for_export,
                                         competitor_summary=comp_summary_for_export)
        st.download_button("Pobierz PDF (raport dla klienta)", data=pdf_bytes,
                            file_name=f"{file_slug}_raport_klienta.pdf",
                            mime="application/pdf")

    st.subheader("5. Google Trends — wykres dla wybranej frazy")
    trends = st.session_state.trends_data
    if trends:
        chosen = st.selectbox("Wybierz frazę", options=sorted(trends.keys()))
        tdf = trends.get(chosen)
        if tdf is not None and not tdf.empty:
            tdf_plot = tdf.copy()
            tdf_plot["data"] = pd.to_datetime(tdf_plot["data"], errors="coerce")
            tdf_plot = tdf_plot.dropna(subset=["data"]).sort_values("data").set_index("data")
            st.line_chart(tdf_plot["trend"])
        else:
            st.info("Brak danych trendu dla tej frazy.")
    else:
        st.info("Brak danych Google Trends — sprawdź klucze API lub spróbuj ponownie.")
elif st.session_state.report_df is not None:
    st.info("Nie udało się pobrać danych. Sprawdź klucze API i spróbuj ponownie.")

# ---------------------------------------------------------------------------
# Analiza konkurencji (opcjonalna, ostatnia sekcja) — same frazy brandowe
# konkurentów, tylko wolumen z DataForSEO (bez Google Trends)
# ---------------------------------------------------------------------------

st.divider()
st.subheader("6. Analiza konkurencji (opcjonalnie)")
st.caption(
    "Wprowadź frazy brandowe konkurentów i przypisz do nich nazwę konkurenta. Sprawdzimy tylko "
    "wolumen wyszukiwań (aktualny i sprzed roku) z DataForSEO — bez Google Trends."
)

if "competitor_input" not in st.session_state:
    st.session_state.competitor_input = pd.DataFrame(
        [{"Konkurent": "", "Fraza": ""} for _ in range(3)]
    )

competitor_input = st.data_editor(
    st.session_state.competitor_input,
    num_rows="dynamic",
    use_container_width=True,
    key="competitor_editor",
    column_config={
        "Konkurent": st.column_config.TextColumn("Konkurent", help="Nazwa marki konkurencji"),
        "Fraza": st.column_config.TextColumn("Fraza", help="Fraza brandowa powiązana z tym konkurentem"),
    },
)

run_competitors = st.button("Pobierz dane o konkurencji", disabled=not dfs_login or not dfs_password)

if "competitor_df" not in st.session_state:
    st.session_state.competitor_df = None
if "competitor_summary" not in st.session_state:
    st.session_state.competitor_summary = None

if run_competitors:
    valid_rows = competitor_input.dropna(subset=["Konkurent", "Fraza"])
    valid_rows = valid_rows[(valid_rows["Konkurent"].str.strip() != "") & (valid_rows["Fraza"].str.strip() != "")]
    if valid_rows.empty:
        st.warning("Uzupełnij przynajmniej jedną parę Konkurent + Fraza.")
    elif not dfs_login or not dfs_password:
        st.error("Podaj login i hasło do DataForSEO w panelu bocznym.")
    else:
        mapping_df = valid_rows.rename(columns={"Konkurent": "konkurent", "Fraza": "fraza"}).copy()
        mapping_df["fraza"] = mapping_df["fraza"].str.strip()
        mapping_df["konkurent"] = mapping_df["konkurent"].str.strip()
        unique_phrases = mapping_df["fraza"].drop_duplicates().tolist()

        with st.spinner(f"Pobieram wolumen dla {len(unique_phrases)} fraz konkurencji (DataForSEO)..."):
            try:
                comp_vol_df = dfs_search_volume(dfs_login, dfs_password, unique_phrases,
                                                 int(location_code), language_code)
            except Exception as e:
                st.error(f"Błąd DataForSEO (search_volume, konkurencja): {e}")
                comp_vol_df = pd.DataFrame()

        if comp_vol_df.empty:
            st.warning("Nie udało się pobrać danych o wolumenie dla podanych fraz konkurencji.")
        else:
            competitor_df = mapping_df.merge(
                comp_vol_df[["fraza", "wolumen_aktualny", "wolumen_rok_temu", "zmiana_%"]],
                on="fraza", how="left",
            )
            st.session_state.competitor_df = competitor_df
            st.session_state.competitor_summary = compute_competitor_summary(competitor_df)

if st.session_state.competitor_df is not None and not st.session_state.competitor_df.empty:
    comp_df = st.session_state.competitor_df
    comp_summary = st.session_state.competitor_summary

    st.markdown("**Podsumowanie wg konkurenta**")
    st.dataframe(comp_summary, hide_index=True, use_container_width=True)

    if comp_summary is not None and not comp_summary.empty:
        st.bar_chart(comp_summary.set_index("konkurent")["wolumen_teraz"])

    st.markdown("**Szczegóły — frazy wg konkurenta**")
    st.dataframe(
        comp_df[["konkurent", "fraza", "wolumen_aktualny", "wolumen_rok_temu", "zmiana_%"]],
        hide_index=True, use_container_width=True,
    )

st.divider()
st.caption(
    "Uwaga: dane DataForSEO oparte są o Google Ads (wolumen) i Google Trends (popularność w skali 0-100). "
    "Wolumen 'rok temu' pochodzi z historii miesięcznej (monthly_searches). "
    "Integracja Senuto jest szkieletowa — endpoint keyword_database wymaga potwierdzenia zgodnie z Twoim planem API."
)
